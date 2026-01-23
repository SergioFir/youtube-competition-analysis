"""
Import seed data from VidIQ Excel export.

Reads channels and their VPH (Views Per Hour) metrics,
then creates channels and seed baselines in the database.

Usage:
    python scripts/import_seed_data.py
"""
import sys
sys.path.insert(0, '.')

from openpyxl import load_workbook
from loguru import logger

from src.database.connection import get_client
from src.database.channels import add_channel, get_channel
from src.youtube.api import YouTubeAPI

# Configure logging
logger.remove()
logger.add(sys.stderr, format="{time:HH:mm:ss} | {level} | {message}", level="INFO")

EXCEL_PATH = r"C:\Users\firul\Downloads\Video metrics.xlsx"


def vph_to_window_views(vph: float) -> dict:
    """
    Convert VPH (Views Per Hour) to estimated views at each window.

    This is a rough estimate assuming linear growth, which isn't accurate
    but provides a starting baseline until real data is collected.
    """
    return {
        "1h": int(vph * 1),
        "6h": int(vph * 6),
        "24h": int(vph * 24),
        "48h": int(vph * 48),
    }


def import_channel(api: YouTubeAPI, client, url: str, vph: float) -> bool:
    """
    Import a single channel with its seed baseline.

    Returns True if successful, False otherwise.
    """
    # Resolve channel URL to get details
    logger.info(f"Resolving: {url}")
    info = api.resolve_channel_url(url)

    if not info:
        logger.error(f"  Could not resolve channel: {url}")
        return False

    channel_id = info["channel_id"]
    channel_name = info["channel_name"]

    # Check if already exists
    existing = get_channel(channel_id)
    if existing:
        logger.info(f"  Channel already exists: {channel_name}")
    else:
        # Add channel
        result = add_channel(
            channel_id=channel_id,
            channel_name=channel_name,
            subscriber_count=info["subscriber_count"],
            total_videos=info["total_videos"],
        )
        if result:
            logger.info(f"  Added channel: {channel_name} ({info['subscriber_count']:,} subs)")
        else:
            logger.error(f"  Failed to add channel: {channel_name}")
            return False

    # Create seed baselines (for Long-form videos only, assuming VidIQ data is for long videos)
    window_views = vph_to_window_views(vph)

    for window_type, views in window_views.items():
        try:
            client.table("channel_baselines").upsert({
                "channel_id": channel_id,
                "is_short": False,  # Assuming VidIQ data is for long-form
                "window_type": window_type,
                "median_views": views,
                "median_likes": None,  # We don't have this data
                "median_comments": None,
                "sample_size": 10,  # VidIQ uses last 10 videos
                "source": "manual",
                "updated_at": "now()",
            }).execute()
        except Exception as e:
            logger.error(f"  Error creating baseline for {window_type}: {e}")

    logger.info(f"  Baselines: 1h={window_views['1h']:,}, 24h={window_views['24h']:,}, 48h={window_views['48h']:,}")
    return True


def main():
    print("\n=== Importing Seed Data ===\n")

    # Load Excel file
    wb = load_workbook(EXCEL_PATH)
    sheet = wb["Sheet1"]

    # Skip header row
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    print(f"Found {len(rows)} channels to import\n")

    api = YouTubeAPI()
    client = get_client()

    success = 0
    failed = 0

    for url, vph in rows:
        if url and vph:
            if import_channel(api, client, url, float(vph)):
                success += 1
            else:
                failed += 1
        print()

    print(f"\n=== Import Complete ===")
    print(f"Success: {success}")
    print(f"Failed: {failed}")


if __name__ == "__main__":
    main()
