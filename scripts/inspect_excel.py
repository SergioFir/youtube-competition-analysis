"""
Inspect the Excel file to understand its structure.
"""
import sys
sys.path.insert(0, '.')

from openpyxl import load_workbook

filepath = r"C:\Users\firul\Downloads\Video metrics.xlsx"

wb = load_workbook(filepath)
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    print(f"Rows: {sheet.max_row}, Columns: {sheet.max_column}\n")

    # Print first 10 rows to understand structure
    print("First 10 rows:")
    for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        print(f"  Row {i}: {row}")
    print()
